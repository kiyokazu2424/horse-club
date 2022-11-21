## 1日分の処理
# wishからそれぞれの分にだれが入れるかをkaituke_listまとめる

# daysからどの日にちにどう入るかの配列on_dutyを定める、これを最後に出力する

# point_file(csv)からpoint_dataを取得し、だれを入れるかを決める

# 最後にpoint_fileにポイントを加えて保存する

## 全体の処理(実行するメイン関数)

# 上の1日分の処理を必要な日数分for文で行う、1日ごとにちゃんとcsvファイルの回数が増えてることが重要

import csv
import calendar , jpholiday
import openpyxl
from openpyxl.styles import PatternFill

## エクセルファイルを作成する関数
def make_excel_file(file_name,sheet_name,year,month,member_list):
# エクセルファイルの作成
  wb = openpyxl.Workbook()
# ファイルの作成時に同時に作成される初期シート
  defo_sheet = wb['Sheet']
# 指定されたシート名に沿ってシートの作成、変数への格納
  ws = wb.create_sheet(f"{sheet_name}",0)
# 初期シートの削除
  wb.remove(defo_sheet)

  c = calendar.Calendar(firstweekday=0)
# カレンダー作成のための月の日にち、曜日、休日のリスト作成
  day_list = c.monthdays2calendar(year,month)
  youbi_list = ['月','火','水','木','金','土','日']
  holiday_list = jpholiday.month_holidays(year,month)

# 青のセルカラー
  fill_a = PatternFill(patternType='solid', fgColor='4169e1')
# 赤のセルカラー  
  fill_b = PatternFill(patternType='solid', fgColor='ff0000')
# 月の記入
  ws.cell(row=1,column=2).value = f'{month}月'
# メンバーを記入
  for mem in range(len(member_list)):
    ws.cell(row=3+mem,column=1).value = member_list[mem]
# 日付を記入
  day_count = 3
  for lis in range(len(day_list)):
    for day,youbi in day_list[lis]:
      # 日付が０でなければ日付と曜日を記入
      if day != 0:
        ws.cell(row=1,column=day_count).value = day
        ws.cell(row=2,column=day_count).value = youbi_list[youbi]
        # 土曜日なら青色に
        if youbi == 5:
          for mem in range(len(member_list) + 2):
            ws.cell(row=mem+1,column=day_count).fill = fill_a
        # 日曜日なら赤色に    
        if youbi == 6:
          for mem in range(len(member_list) + 2):
            ws.cell(row=mem+1,column=day_count).fill = fill_b
        # 祝日であるかを確認して祝日なら赤色に(土曜が祝日ならば上書きされて青ではなく赤になる)
        for holiday in range(len(holiday_list)):
          if day == holiday_list[holiday][0].day:
            for mem in range(len(member_list) + 2):
              ws.cell(row=mem+1,column=day_count).fill = fill_b
        # 日付を追加後、次の日付を作る前にカウントを1増やすことで隣の欄に記入できるようにする
        day_count += 1    
# エクセルファイルをセーブ
  wb.save(f'{file_name}.xlsx')




## エクセルファイルに指定のシートを作成する関数(基本的に上の関数同様)
def make_excel_calendar(excel_file,sheet_name,year,month,member_list):
  # 指定エクセルファイルの読み取り
  wb = openpyxl.load_workbook(f"{excel_file}.xlsx")

  #指定されたシート名に沿ってシートの作成、変数への格納
  ws = wb.create_sheet(f"{sheet_name}",month-1)

  c = calendar.Calendar(firstweekday=0)
  day_list = c.monthdays2calendar(year,month)
  youbi_list = ['月','火','水','木','金','土','日']
  holiday_list = jpholiday.month_holidays(year,month)


# 青のセルカラー
  fill_a = PatternFill(patternType='solid', fgColor='4169e1')
# 赤のセルカラー  
  fill_b = PatternFill(patternType='solid', fgColor='ff0000')
# 月の記入
  ws.cell(row=1,column=2).value = f'{month}月'
# メンバーを記入
  for mem in range(len(member_list)):
    ws.cell(row=3+mem,column=1).value = member_list[mem]
# 日付を記入
  day_count = 3
  for lis in range(len(day_list)):
    for day,youbi in day_list[lis]:
      # 日付が０でなければ日付と曜日を記入
      if day != 0:
        ws.cell(row=1,column=day_count).value = day
        ws.cell(row=2,column=day_count).value = youbi_list[youbi]
        # 土曜日なら青色に
        if youbi == 5:
          for mem in range(len(member_list) + 2):
            ws.cell(row=mem+1,column=day_count).fill = fill_a
        # 日曜日なら赤色に    
        if youbi == 6:
          for mem in range(len(member_list) + 2):
            ws.cell(row=mem+1,column=day_count).fill = fill_b
        # 祝日であるかを確認して祝日なら赤色に   
        for holiday in range(len(holiday_list)):
          if day == holiday_list[holiday][0].day:
            for mem in range(len(member_list) + 2):
              ws.cell(row=mem+1,column=day_count).fill = fill_b
        # 日付を追加後、次の日付を作る前にカウントを1増やすことで隣の欄に記入できるようにする
        day_count += 1
# エクセルファイルに上書き
  wb.save(f'{excel_file}.xlsx')




## 1日分の飼付の決定処理
def one_day_on_duty(point_file,wish,day,excel_file,year,member_list):
  ##csvの処理
  #csvを読み込み、配列point_dataに情報を格納する
  csv_file = open(f"/Users/kiyo/Desktop/馬術部_開発/csv_for_tame/{point_file}.csv", "r", encoding="utf_8", errors="", newline="" )
  #読み込んだファイルをリスト形式で返す(csv)
  f_lis = csv.reader(csv_file, delimiter=",", doublequote=True, lineterminator="\r\n", quotechar='"', skipinitialspace=True)
  header_lis = next(f_lis)

  #csvのデータとそれをソートするためのリスト、リストのリストで各リストが一つのデータ
  point_data = []

  #リストに各行のデータを格納、header_lisはcsvファイル一行目の項目のリスト
  point_data.append(header_lis)
  for row in f_lis:
    point_data.append(row)

  csv_file.close()
  # print(point_data)

## エクセルファイルの読み込み、シートの追加

# エクセルファイルの読み込み
  wb = openpyxl.load_workbook(f"{excel_file}.xlsx")
  
# 決めたい飼付の日付の月に対応するシートが存在するか調べる
# 今決めている対象の日にちの月を取り出す
  month = day[0].split('/')[0]

# シートにある月をリストにする
  sheet_list = []
  for i in range(len(wb.sheetnames)):
    sheet_month = wb.sheetnames[i].split('月')[0]
    sheet_list.append(sheet_month)
# シートに無かったら作る
  if month not in sheet_list:
    make_excel_calendar(f'{excel_file}',f'{month}月',year,int(month),member_list)

### 総合回数->各posの回数で判断
## 朝〜夜で入れる人の名前をそれぞれピックアップする
  kaituke_list = [[],[],[],[]]
  
  # 以下基本的にposは朝、昼、夕、夜を決める数値
  # 朝〜夜のfor文
  for pos in range(len(kaituke_list)):
    # 各メンバーの希望
    for k in range(len(wish)):
      # wishの値が0(可能)であればその人の名前を追加していく
      if wish[k][pos+3] == 0:
        kaituke_list[pos].append(wish[k][2])

  # print(kaituke_list)
  
  # point_listに飼付に入れる人のid,総合ポイント,posのポイントをkaituke_listを参照に朝〜夜それぞれまとめる
  # 人のid,総合ポイント,posポイントの組みの配列がをそれぞれ追加している
  point_list = [[],[],[],[]]
  for pos in range(len(kaituke_list)):
    for member in kaituke_list[pos]:
      for k in range(len(point_data)):
        if member == point_data[k][2]:
          point = []
          point.append(point_data[k][0])
          point.append(point_data[k][7])
          point.append(point_data[k][pos+3])
          point_list[pos].append(point)

  # print(point_list)
  
  # 総合ポイントでソート後、一番少ないポイントの人が複数いた場合、posポイントで決める
  # デフォルト（朝〜夜全て必要な場合）
  if len(day) == 1:
    pos_list = ['朝','昼','夕','夜']
    update_list = []
  # 朝〜夜
    for pos in range(len(pos_list)):

      #print(pos_list[pos])
      search_all_point = []
      # 総合ポイントの最小を割り出す
      for k in range(len(point_list[pos])):
        if point_list[pos][k][1] not in search_all_point:
          search_all_point.append(point_list[pos][k][1])
      # 最小値の格納、point_dataには文字列として数値が入ってるので一旦数値にして比較しなければならない
      # print(search_all_point)
      # print(type(search_all_point[0]))
      search_all_point_sorted = [int(n) for n in search_all_point]
      minmum = min(search_all_point_sorted)

      hit_person = []
      # 総合ポイントが一番少なかった人を割り出してリストに入れる
      for k in range(len(point_list[pos])):
        if point_list[pos][k][1] == str(minmum):
          hit_person.append(point_list[pos][k])

      # 一番少なかった人が一人しかいない場合
      if len(hit_person) == 1:
        # point_dataのidと照らし合わせて見つける
        for n in range(len(point_data)):
          if point_data[n][0] == hit_person[0][0]:
            member = point_data[n][2]
            day.append(f'{pos_list[pos]}:{member}')
            
            update_list.append(pos_list[pos])
            update_list.append(member)

      # 複数同じポイントの人がいる場合posポイントで決める
      else:
        search_pos_point = []
        # posポイントの最小を割り出す
        for k in range(len(hit_person)):
          if hit_person[k][2] not in search_pos_point:
            search_pos_point.append(hit_person[k][2])
        # 最小値の格納、point_dataには文字列として数値が入ってるので一旦数値にして比較しなければならない
        search_pos_point_sorted = [int(n) for n in search_pos_point]
        minmum2 = min(search_pos_point_sorted)

        # 飼付が決まった人のidを割り出す
        for k in range(len(hit_person)):
          if hit_person[k][2] == str(minmum2):
            person_id = hit_person[k][0]
            break

        # point_dataのidと上で決めたperson_idを照らし合わせて見つける
        for n in range(len(point_data)):
          if point_data[n][0] == person_id:
            member = point_data[n][2]
            day.append(f'{pos_list[pos]}:{member}')
            update_list.append(pos_list[pos])
            update_list.append(member)



  # カスタムとして必要な時間（朝、昼とか）を入力している場合
  else:
    pos_list = []
    update_list = []
    sorted_point_list = []
    # 必要とされている飼付のタイミングだけにする
    if '朝' in day:
      pos_list.append('朝')
      sorted_point_list.append(point_list[0])
      day.remove('朝')
    if '昼' in day:
      pos_list.append('昼')
      sorted_point_list.append(point_list[1])
      day.remove('昼')
    if '夕' in day:
      pos_list.append('夕')
      sorted_point_list.append(point_list[2])
      day.remove('夕')
    if '夜' in day:
      pos_list.append('夜')
      sorted_point_list.append(point_list[3])
      day.remove('夜')

    for pos in range(len(pos_list)):
      search_point = []
      for k in range(len(sorted_point_list[pos])):
        if sorted_point_list[pos][k][1] not in search_point:
          search_point.append(sorted_point_list[pos][k][1])
      minmum = min(search_point)
      for s in range(len(sorted_point_list[pos])):
        if sorted_point_list[pos][s][1] == minmum:
          for n in range(len(point_data)):
            if point_data[n][0] == sorted_point_list[pos][s][0]:
              member = point_data[n][2]
          day.append(f'{pos_list[pos]}:{member}')
          update_list.append(pos_list[pos])
          update_list.append(member)
          break

  ## csvファイルをupdate_listの情報をもとに更新する

  # 初めに読み込んだデータを更新する
  
  # update_listは必ずposと人の二つ人組で朝〜夜全てがないパターンもあるので
  for i in range(int(len(update_list)/2)):
    for k in range(len(point_data)):
      # かく飼付入ってる人の名前が入ったデータがあれば
      if update_list[2*i+1] in point_data[k]:
        # 総合ポイントを増やす
        point_data[k][7] = int(point_data[k][7]) + 1
        # posポイントを増やす
        for n in range(len(point_data[0])):
          if update_list[2*i] == point_data[0][n]:
            point_data[k][n] = int(point_data[k][n]) + 1
  # csvファイルを更新したデータにする(csvファイルに反映する)
  csv_file = open(f"/Users/kiyo/Desktop/馬術部_開発/csv_for_tame/{point_file}.csv", "w", encoding="utf_8")
  writer = csv.writer(csv_file)
  writer.writerows(point_data)


  # エクセルファイルに当番を書き込む
  month = int(day[0].split('/')[0])
  day_num = int(day[0].split('/')[1])
  ws = wb[f'{month}月']
  # 該当する日付の列の文字を格納
  col_let = ws.cell(row=1,column=day_num+2).column_letter
  # 各飼付の分を記入
  
  for i in range(int(len(update_list)/2)):
    for row in range(ws.max_row):

      if ws[f'A{row+1}'].value == update_list[2*i+1]:
        if ws[f'{col_let}{row+1}'].value == None:
          ws[f'{col_let}{row+1}'].value = update_list[2*i]
        else:
          ws[f'{col_let}{row+1}'].value = ws[f'{col_let}{row+1}'].value + ' ' + update_list[2*i]
  wb.save(f'{excel_file}.xlsx')

  print(update_list)
  print(day)




## 上の処理を指定した日数分だけ行うメイン関数
def on_duty(point_file,wish_list,day_list,excel_file,year,member_list,mode):
  if mode == 'new':
    make_excel_file(f'飼付{year}','1月',year,1,member_list)
  for d in range(len(day_list)):
    wish = wish_list[d]
    day = day_list[d]
    one_day_on_duty(point_file,wish,day,excel_file,year,member_list)




## 平日の処理、基本学期中は固定のため月曜日〜金曜日までをweekday_infoに入力しておく
def weekday_on_duty(point_file,weekday_info,excel_file):
  ##csvの処理
  #csvを読み込む
  csv_file = open(f"/Users/kiyo/Desktop/馬術部_開発/csv_for_tame/{point_file}.csv", "r", encoding="utf_8", errors="", newline="" )
  #読み込んだファイルをリスト形式で返す(csv)
  f_lis = csv.reader(csv_file, delimiter=",", doublequote=True, lineterminator="\r\n", quotechar='"', skipinitialspace=True)
  header_lis = next(f_lis)

  #csvのデータとそれをソートするためのリスト、リストのリストで各リストが一つのデータ
  point_data = []

  #リストに各行のデータを格納、header_lisはcsvファイル一行目の項目のリスト
  point_data.append(header_lis)
  for row in f_lis:
    point_data.append(row)

  csv_file.close()

  # 各曜日に対して昼、夜のポイント、合計のポイントの加算処理
  for weekday in range(len(weekday_info)):
    # 昼の処理
    for member in range(len(point_data)):
      if point_data[member][2] == weekday_info[weekday][2]:
        # 昼のポイント追加
        point_data[member][4] = int(point_data[member][4]) + 1
        # 総合のポイント追加
        point_data[member][7] = int(point_data[member][7]) + 1
    # 夜の処理    
    for member in range(len(point_data)):
      if point_data[member][2] == weekday_info[weekday][4]:
        # 夜のポイント追加
        point_data[member][6] = int(point_data[member][6]) + 1
        # 総合のポイント追加
        point_data[member][7] = int(point_data[member][7]) + 1

  # csvファイルを更新したデータにする
  csv_file = open(f"/Users/kiyo/Desktop/馬術部_開発/csv/{point_file}.csv", "w", encoding="utf_8")

  writer = csv.writer(csv_file)
  writer.writerows(point_data)

  # 月を跨ぐ処理がめんどい、weekdayの最後にはじまりと終わりを入力してもらい自動取得にするか？
  # エクセルファイルに当番を書き込む
  # wb = openpyxl.load_workbook(f"{excel_file}.xlsx")
  # month = int(day[0].split('/')[0])
  # day_num = int(day[0].split('/')[1])
  # ws = wb[f'{month}月']
  # # 該当する日付の列の文字を格納
  # col_let = ws.cell(row=1,column=day_num+2).column_letter
  # # 各飼付の分を記入
  
  # for i in range(int(len(update_list)/2)):
  #   for row in range(ws.max_row):

  #     if ws[f'A{row+1}'].value == update_list[2*i+1]:
  #       if ws[f'{col_let}{row+1}'].value == None:
  #         ws[f'{col_let}{row+1}'].value = update_list[2*i]
  #       else:
  #         ws[f'{col_let}{row+1}'].value = ws[f'{col_let}{row+1}'].value + ' ' + update_list[2*i]
  # wb.save(f'{excel_file}.xlsx')


## 関数１ make_excel_fileについて

""" 引数1 ファイル名の指定
引数2 シートの名前の指定、基本的にn月
引数3,4 年と月 シートの横軸を作成するために参照する
引数5 メンバーの一覧 縦軸を作成するときに参照する """

## 実行
member_list = ['Aさん','Bさん','Cさん','Dさん','Eさん','Fさん','Gさん','Hさん']
# make_excel_file('飼付2022','1月',2022,1,member_list)


## 関数２ make_excel_calenderについて

"""
引数1 追加する対象のエクセルファイル名
引数2 シートの名前の指定、基本的にn月
引数3、4 年と月 シートの横軸を作成するために参照する
引数5 メンバーの一覧 縦軸を作成するときに参照する
 """

## 実行
# make_excel_calendar('飼付2022','4月',2022,4,member_list)


## 関数3 one_day_on_dutyについて

"""
引数1 ポイントファイル 飼付のポイントが書かれているファイルのファイル名
引数2 飼付希望 何らかの形で['名前',0 or 1,0 or 1,0 or 1,0 or 1] (0ができる1ができない)のという配列の配列を受け取る、今は手入力
テンプレート [0,3,'Aさん',1,1,1,1],[1,3,'Bさん',1,1,1,1],[2,3,'Cさん',1,1,1,1],[3,2,'Dさん',1,1,1,1],[4,2,'Eさん',1,1,1,1],[5,1,'Fさん',1,1,1,1],[6,1,'Gさん',1,1,1,1],[7,1,'Hさん',1,1,1,1]
引数3 飼付を決める日数と時間（朝、昼とか）の配列 ['日付']や['日付','昼','夜'] (何も指定がなければ朝〜夜全てに入れる)
引数4 書き込む対象のエクセルファイルの名前
引数5 飼付を行う年
単純に一日の飼付とかその変更で使ってもいいかも
"""
wish = [[0,3,'Aさん',0,1,1,0],[1,3,'Bさん',0,1,0,1],[2,3,'Cさん',0,0,0,0],[3,2,'Dさん',1,1,0,0],[4,2,'Eさん',0,0,1,1],[5,1,'Fさん',0,0,0,1],[6,1,'Gさん',0,1,0,0],[7,1,'Hさん',0,1,1,0]]
day = ['1/22']

## 実行
# one_day_on_duty('kaituke_point',wish,day,'飼付2022',2022)


## 関数4 on_dutyについて

"""
引数1 ポイントファイル 上の関数同様
引数2 飼付希望 上の引数2の日数分の配列
引数3 入る必要がある日 上の引数3の日数分の配列
引数4 書き込む対象のエクセルファイルの名前
引数5 飼付を行う年
引数6 部員のリスト、エクセルファイルやシートの作成時に使用する
引数7 新しくエクセルファイルを作るなら'new'とし今あるのに追加したいなら他の文字列を打ち込んどけばよい、基本は後者で使う
"""
wish_list = [[[0,3,'Aさん',0,1,1,0],[1,3,'Bさん',0,1,0,1],[2,3,'Cさん',0,1,1,0],[3,2,'Dさん',1,1,0,0],[4,2,'Eさん',0,1,1,1],[5,1,'Fさん',0,0,0,1],[6,1,'Gさん',0,1,0,0],[7,1,'Hさん',0,1,1,0]],
            [[0,3,'Aさん',1,1,1,0],[1,3,'Bさん',1,1,0,1],[2,3,'Cさん',0,1,1,0],[3,2,'Dさん',1,0,0,0],[4,2,'Eさん',0,0,1,1],[5,1,'Fさん',0,1,0,1],[6,1,'Gさん',0,1,0,0],[7,1,'Hさん',0,1,1,0]]]
day_list = [['1/29'],['1/30']]

## 実行
member_list = ['Aさん','Bさん','Cさん','Dさん','Eさん','Fさん','Gさん','Hさん']
on_duty('kaituke_point',wish_list,day_list,'飼付2022',2022,member_list,'new')


## 関数5 weekday_on_dutyについて

"""
引数1 ポイントファイル 上の関数同様
引数2 平日の当番情報 各曜日分の配列
引数3 追加対象のエクセルファイル
"""
weekday_info = [['月曜日','昼','Aさん','夜','Aさん'],['火曜日','昼','Aさん','夜','Aさん'],['水曜日','昼','Aさん','夜','Aさん'],['木曜日','昼','Bさん','夜','Bさん'],['金曜日','昼','Bさん','夜','Bさん'],1/3,1/7]

## 実行
# 平日分を実行するときは下のコメントアウトを外す、ずっと機能させっぱなしだと必要以上に入ったことになってしまう
# weekday_on_duty('kaituke_point',weekday_info)


"""
使用ケース1 長期休みではない時の土日の飼付 関数4を使用
使用ケース2 長期休みではない時の平日の飼付 関数5を使用
使用ケース3 長期休みの時の飼付 関数4を使用 
どのケースにおいても月を跨ぐ可能性があり、跨ぐか否かで分岐、跨ぐ場合（指定した月のエクセル表が存在しない場合）新しくその月のエクセル表が増える仕様、年を越す場合に限り新しくエクセルファイル自体を作成する
"""



""" 懸念ポイント
だれも入れない場所があるとバグる(治せはする)
飼付の入れる日を入力するのがめんどくさい
土日だけを決める場合、朝〜夜の回数だけで決めると全体の回数にばらつきが出てしまう（飼付の合計回数を考慮するコードを挟むべき）
この人は土日入るのが難しいから〜などの調整は手作業
急な当番交代も手作業
飼付がないところも全て0にしないといけない、をの場合無視できるようにしている
平日入ってる人の優先度を下げたりするのはcsvファイルの枠を追加することで操作可能だと思う
そもそも全部入れない人はリストに入れなくていいかも!! """




""" 改善すべき点
平日は別ファイルで追加、修正できるようにする -> 毎週実行するのがだるいなら先に何回かわかってる分だけ実行するののもアリ？
少ない場合の警告→少し難しいかも
月毎にエクセルを作成、土日と休日は色付け
カレンダーを自動で取ってくるコードを設けたい
csvファイルを一回の買い付けごとにupdateするか、その日ごとにupdateするかによって挙動が変わる、どちらにしろ総合的には偏ることがないように出来ている
日付ごとにしたほうが1日に固まるのでかえってシステムとして良いかも、とりあえずこのままにしとく
"""

